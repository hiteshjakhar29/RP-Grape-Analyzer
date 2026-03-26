"""exporter.py — Excel export for grape measurement results."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


_HEADERS = [
    "Grape ID", "Group", "Ratio", "Row", "Day", "Area (px)",
    "Mean R", "Std R", "Mean G", "Std G", "Mean B", "Std B",
    "Mean L*", "Std L*", "Mean a*", "Std a*", "Mean b*", "Std b*",
    "Hue (°)", "Std H", "Saturation", "Std S", "Brightness", "Std Br",
    "Eccentricity",
]
_KEYS = [
    "grape_id", "group", "ratio", "grape_row", "day", "area_px",
    "mean_R", "std_R", "mean_G", "std_G", "mean_B", "std_B",
    "mean_L", "std_L", "mean_a", "std_a", "mean_b", "std_b",
    "mean_H", "std_H", "mean_S", "std_S", "mean_Br", "std_Br",
    "eccentricity",
]

_HDR_FILL  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HDR_FONT  = Font(color="FFFFFF", bold=True, size=11)
_COAT_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
_CTRL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def export_to_excel(results: list, session_day: int, output_path: str) -> None:
    """
    Write results list to an Excel file.

    Parameters
    ----------
    results      : list of dicts from color_engine.measure_grape
    session_day  : int — study day number (used as sheet name)
    output_path  : str — destination .xlsx path
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Day_{session_day}"

    # Header row
    for col, h in enumerate(_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill      = _HDR_FILL
        cell.font      = _HDR_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Data rows
    for row_i, r in enumerate(results, 2):
        fill = _COAT_FILL if r.get("group") == "Coated" else _CTRL_FILL
        for col, key in enumerate(_KEYS, 1):
            val  = r.get(key, "")
            cell = ws.cell(row=row_i, column=col, value=val)
            cell.fill      = fill
            cell.alignment = Alignment(horizontal="center")

    # Auto-fit column widths
    for col in ws.columns:
        max_w = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_w + 4

    wb.save(output_path)
